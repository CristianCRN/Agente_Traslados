from openpyxl import load_workbook
from database.operaciones import traslado_mes

def generar_reporte(mes,ano):
    try: 
        # PASO 1 DICCIONARIOS DE LOS MESES
        meses = {
            "01": "ENERO", "02": "FEBRERO", "03": "MARZO",
            "04": "ABRIL", "05": "MAYO", "06": "JUNIO",
            "07": "JULIO", "08": "AGOSTO", "09": "SEPTIEMBRE",
            "10": "OCTUBRE", "11": "NOVIEMBRE", "12": "DICIEMBRE"
        }
        # PASO 2 ABRO EL LIBRO INDICADO 
        wb = load_workbook("Formato Planilla de Transporte.xlsx")
        
        # PASO 3 CONDICIONAL SI EXISTE LA HOJA
        nombre_mes = meses[mes]
        nombre_hoja = f"{nombre_mes} {ano}"

        if nombre_hoja in wb.sheetnames:
            hoja = wb[nombre_hoja]
        else:
            hoja = wb.copy_worksheet(wb.active)
            hoja.title = nombre_hoja
        # PASO 4 Y 5 CONSULTAR LA BASE DE DATOS CON Traslados_MES
        filas = traslado_mes(mes, ano)
        fila_actual = 6
        for fila in filas:
            hoja.cell(row=fila_actual, column=2).value = fila[1]
            hoja.cell(row=fila_actual, column=3).value = fila[2]
            hoja.cell(row=fila_actual, column=4).value = fila[3]
            hoja.cell(row=fila_actual, column=5).value = fila[4]
            hoja.cell(row=fila_actual, column=6).value = fila[5]
            hoja.cell(row=fila_actual, column=7).value = fila[6]
            fila_actual += 1
        
        wb.save("Formato Planilla de Transporte.xlsx")
        return "Formato Planilla de Transporte.xlsx"
    except Exception as e:
        print(f"Error escribiendo el excel: {e}")
        return False